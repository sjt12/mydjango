from django import forms
from django.contrib import admin
from django.http import HttpResponse
from taggit.forms import TagField
from taggit_helpers.admin import TaggitStackedInline as TSI, TaggitTabularInline as TTI, TaggitListFilter as TLF
from taggit_labels.widgets import LabelWidget

from core.models import User
from .models import Question, Response

class ContentForm(forms.ModelForm):
	tags = TagField(required=False, widget=LabelWidget)
	
class TagAdmin(TTI):
	form = ContentForm

def export_xlsx(modeladmin, request, queryset):
	import openpyxl
	from openpyxl.utils import get_column_letter
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=responses.xlsx'
	wb = openpyxl.Workbook()
	ws = wb.get_active_sheet()
	ws.title = "Responses"

	row_num = 0

	columns = [
		(u"ResponseID", 15),
		(u"QuestionID", 15),
		(u"QuestionTags", 15),
		(u"UserID", 15),
		(u"Age", 15),
		(u"Education", 15),
		(u"Accepted", 15),
		(u"Correction", 50)
	]

	for col_num in range(len(columns)):
		c = ws.cell(row=row_num + 1, column=col_num + 1)
		c.value = columns[col_num][0]
		#c.style.font.bold = True
		# set column width
		ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

	for obj in queryset:
		row_num += 1
		thetags = ''
		for tag in obj.question.tags.all():
			if thetags == '':
				thetags = tag.slug
			else:
				thetags = thetags + ", " + tag.slug
		row = [
			obj.pk,
			obj.question.id,
			thetags,
			obj.user.id,
			obj.user.age,
			obj.user.education,
			obj.accepted,
			obj.correction,
		]
		for col_num in range(len(row)):
			c = ws.cell(row=row_num + 1, column=col_num + 1)
			c.value = row[col_num]
			#c.style.alignment.wrap_text = True

	wb.save(response)
	return response

export_xlsx.short_description = u"Export XLSX"

class ResponseAdmin(admin.ModelAdmin):
	list_display = ('question', 'tag_list', 'user', 'accepted')
	actions = [export_xlsx]
	list_filter = ['question__tags']
	
	def get_queryset(self, request):
		return super(ResponseAdmin, self).get_queryset(request).prefetch_related('question')
		
	def tag_list(self, obj):
		return u", ".join(o.name for o in obj.question.tags.all())
			
class QuestionAdmin(admin.ModelAdmin):
	fields = ['title', 'content']
	inlines = [TTI]
	list_filter = [TLF]
	list_display = ('id', 'title', 'tag_list', 'responses', 'yeas', 'nays')
	ordering = ['id']
	
	def get_queryset(self, request):
		return super(QuestionAdmin, self).get_queryset(request).prefetch_related('tags')
		
	def tag_list(self, obj):
		return u", ".join(o.name for o in obj.tags.all())

admin.site.register(Question, QuestionAdmin)
admin.site.register(Response, ResponseAdmin)			
